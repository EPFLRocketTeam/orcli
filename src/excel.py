#format simulation data to Excel

#used for writing results to Excel
import openpyxl
#used for taking arguments during program call
import sys

if(sys.argv[1] == "1"):
#if no workbook exists (if this is the 1st simulation)
    #create an Excel workbook as "wb"
    wb = openpyxl.Workbook()
else:
    #open existing Excel workbook as "wb"
    wb = openpyxl.load_workbook('results.xlsx')

name = "Simulation " + sys.argv[1]  #name of simulation

#search "rocket.ork" for simulation results
f = open("rocket.ork")
#for loop reads "rocket.ork" line by line
for x in f:
    #if the line contains "simulationend"
    if(x.find("simulationend") >= 0):
        break
#create a new sheet in wb as "ws"
ws = wb.create_sheet(name)
labels = ["Time","Altitude","Vertical velocity","Vertical acceleration","Total velocity","Total acceleration","Position East of launch","Position North of launch","Lateral distance","Lateral direction","Lateral velocity","Lateral acceleration","Latitude","Longitude","Gravitational acceleration","Angle of attack","Roll rate","Pitch rate","Yaw rate","Mass","Motor mass","Longitudinal moment of inertia","Rotational moment of inertia","CP location","CG location","Stability margin calibers","Mach number","Reynolds number","Thrust","Drag force","Drag coefficient","Axial drag coefficient","Friction drag coefficient","Pressure drag coefficient","Base drag coefficient","Normal force coefficient","Pitch moment coefficient","Yaw moment coefficient","Side force coefficient","Roll moment coefficient","Roll forcing coefficient","Roll damping coefficient","Pitch damping coefficient","Coriolis acceleration","Reference length","Reference area","Vertical orientation (zenith)","Lateral orientation (azimuth)","Wind velocity","Air temperature","Air pressure","Speed of sound","Simulation time step","Computation time"]
for i in range(len(labels)):
    #write labels[i] to row 1, col i
    ws.cell(1,i+1,labels[i])
k = 2  #row of cell to be written to
for x in f:
    #if the line no longer contains simulation data
    if(x.find("datapoint") < 0):
        break
    y = ""  #string to store value to be written to Excel
    j = 1  #column of cell to be written to
    for i in range(21,len(x)):
        if(x[i] == '<'):  #end of data
            #write latest value to Excel
            ws.cell(k,j,y)
            break
        if(x[i] == ','):
        #comma marks separation between 2 values
            #write latest value to Excel
            ws.cell(k,j,y)
            j += 1  #increment column number
            y = ""  #reset y
            continue  #skip the comma
        y += x[i]  #add current character to y
    k += 1  #increment row number
f.close()
wb.save('results.xlsx')
